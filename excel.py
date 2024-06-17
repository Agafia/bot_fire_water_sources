import pandas as pd
from openpyxl import load_workbook
from pyproj import Transformer  # Преобразование координат между проекциями

# wb = load_workbook('41.xlsx')
# ws = wb['Справочник_ИНППВ']
# print(ws.tables.items())
#
# dfs = []
# for table_name, value in ws.tables.items():
#     table = ws[value]
#     for row in table:
#         for cell in row:
#             print(cell.value, cell)
    # header, *body = [[cell.value for cell in row]
    #                   for row in table]
    # print(type(body), type(header))
    # print(header)
    # df = pd.DataFrame(body, columns=header)
    # dfs.append(df)
# print(dfs)

data = pd.read_excel('41.xlsx')
headers_int = ['ИД_хоз_субъекта', 'ИД_вид_ППВ', 'ИД_исп_ППВ', 'ИД_зоны_части', 'ИД_верхего_МО', 'ИД_нижнего_МО',
               'ИД_границ_НП']
headers_str = ['Поселение', 'Улица', 'Дом', 'Вид_ВИ', 'Номер', 'Характеристика', 'Исполнение', 'Способ_обогрева',
               'Указатель_место', 'Указатель_ГОСТ', 'Пирамида', 'Ориентир', 'Состояние', 'Дефект_описание',
               'Водоотдача_сети', 'Регистрация_повод', 'Исключение_повод']
headers_date = ['Дефект_выявлен', 'Дефект_устранён', 'Дата_испытания', 'Регистрация_дата', 'Исключение_дата']
headers_geom = ['Широта', 'Долгота', ]


fields_dict = {}
fields_int = {}
fields_geom = {}
for ind in data.index:

    for column in data:
        if column in headers_int:
            if pd.notnull(data[column][ind]):
                fields_int[column] = int(data[column][ind])
        elif column in headers_str:
            if pd.notnull(data[column][ind]):
                fields_dict[column] = str(data[column][ind])
        elif column in headers_date:

            if pd.notnull(data[column][ind]):
                field_value = pd.to_datetime(data[column][ind])
                fields_date = {'year': "{:02d}".format(field_value.year),
                               'month': "{:02d}".format(field_value.month),
                               'day': "{:02d}".format(field_value.day)}
                fields_dict[column] = fields_date

                # print(data[column][ind], fields_date)
        elif column == 'Широта':
            fields_geom['lat'] = float(data[column][ind])
        elif column == 'Долгота':
            fields_geom['lon'] = float(data[column][ind])

    transformer = Transformer.from_crs("EPSG:4326", "EPSG:3857")
    sm = transformer.transform(fields_geom['lat'], fields_geom['lon'])
    geom = f'POINT({str(sm[0])} {str(sm[1])})'

    print(fields_dict)
    print(fields_int)
    print(fields_geom, geom)
    fields_int.clear()
    fields_dict.clear()
    fields_geom.clear()

    # break


